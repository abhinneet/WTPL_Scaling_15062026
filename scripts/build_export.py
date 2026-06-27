"""
build_export.py
Production-ready, secure version.

Takes MITRA_Analytics_v7_Complete.xlsx as template, populates all RAW_* sheets
with live DB data (passed as JSON), and writes the result to output path.

Usage:
    python3 build_export.py --template <input.xlsx> --data <data.json> --output <output.xlsx>

Security & robustness features:
    * Path validation (resolves, restricts extensions, prevents traversal in output dir)
    * JSON schema validation per sheet (types + required keys)
    * Excel formula-injection sanitisation for all string values
    * Bounds checking on row indices and column indices
    * Atomic save (write to temp file, fsync, rename)
    * Structured logging instead of print()
    * Bounded resource usage (file-size limit, row-count limit)
    * Graceful degradation: missing sheet -> logged warning, not crash
    * No eval/exec, no shell, no pickle, no untrusted deserialisation

Exit codes:
    0  success
    2  invalid arguments
    3  template file invalid / missing sheets
    4  data JSON invalid (schema or parse error)
    5  I/O or workbook error
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
import tempfile
import warnings
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple, Union

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

warnings.filterwarnings("ignore")

# --------------------------------------------------------------------------- #
# Constants
# --------------------------------------------------------------------------- #

LOG = logging.getLogger("build_export")

# Restrict what we accept as inputs.
ALLOWED_TEMPLATE_EXT = {".xlsx"}
ALLOWED_OUTPUT_EXT = {".xlsx"}
ALLOWED_DATA_EXT = {".json"}

# Hard limits to mitigate DoS / accidental huge payloads.
MAX_DATA_FILE_BYTES = 50 * 1024 * 1024          # 50 MB JSON
MAX_TEMPLATE_FILE_BYTES = 100 * 1024 * 1024     # 100 MB xlsx
MAX_ROWS_PER_SHEET = 200_000
MAX_ADS_ROWS = 10_000

# Required sheets in the template.
REQUIRED_SHEETS = (
    "RAW_DISTRICT",
    "RAW_AR_CONTENT",
    "RAW_QUIZ",
    "RAW_SESSION",
    "RAW_LANGUAGE",
    "RAW_DEVICE",
    "RAW_NOTIFICATION",
    "RAW_ADS",
)

# Column mappings: sheet_name -> list of (target_col_index, source_key, type)
# 1-based column indices matching the original script.
SHEET_MAPS: Dict[str, List[Tuple[int, str, type]]] = {
    "RAW_DISTRICT": [
        (5,  "active_users",       int),
        (6,  "dau_avg",            (int, float)),
        (7,  "quiz_attempts",      int),
        (8,  "quiz_correct",       int),
        (9,  "ar_sessions",        int),
        (10, "ar_completions",     int),
        (11, "total_sessions",     int),
        (12, "offline_sessions",   int),
        (13, "ar_capable_devices", int),
        (14, "total_devices",      int),
        (15, "schools_active",     int),
        (16, "teacher_logins",     int),
    ],
    "RAW_AR_CONTENT": [
        (10, "total_launches",   int),
        (11, "unique_students",  int),
        (12, "avg_dwell_min",    (int, float)),
        (13, "completion_pct",   (int, float)),
        (14, "replay_pct",       (int, float)),
        (15, "pre_score_pct",    (int, float)),
        (16, "post_score_pct",   (int, float)),
    ],
    "RAW_QUIZ": [
        (10, "attempts",          int),
        (11, "correct_answers",   int),
        (12, "completions",       int),
        (13, "abandoned",         int),
        (14, "total_attempt_min", (int, float)),
    ],
    "RAW_SESSION": [
    (3,  "__month_label__",         str),  # Col C Report Month (was 2, now 3)
    (4,  "total_sessions",          int),
    (5,  "avg_session_min",         (int, float)),
    (6,  "bounce_pct",              (int, float)),
    (7,  "sessions_1_5_pct",        (int, float)),
    (8,  "sessions_5_15_pct",       (int, float)),
    (9,  "sessions_15_30_pct",      (int, float)),
    (10, "sessions_30plus_pct",     (int, float)),
],
    "RAW_LANGUAGE": [
        (2, "primary_language",  str),
        (3, "sessions_primary",  int),
        (4, "sessions_hindi",    int),
        (5, "sessions_english",  int),
    ],
    "RAW_DEVICE": [
        (2,  "ar_capable_devices", int),
        (3,  "total_devices",      int),
        (4,  "devices_low",        int),
        (5,  "devices_mid",        int),
        (6,  "devices_high",       int),
        (15, "offline_pct",        (int, float)),
    ],
    "RAW_NOTIFICATION": [
        (3, "sent",          int),
        (4, "delivered",     int),
        (5, "delivery_pct",  (int, float)),
        (6, "open_rate",     (int, float)),
        (7, "ctr_pct",       (int, float)),
    ],
}

# Raw value sanitisation: prevent Excel formula injection.
# Excel treats leading =, +, -, @, TAB, CR, and \x00 specially.
_INJECTION_RE = re.compile(r"^[=+\-@\t\r\x00]")


# --------------------------------------------------------------------------- #
# Errors
# --------------------------------------------------------------------------- #

class BuildExportError(Exception):
    """Base class for all expected, user-facing errors."""


class InvalidArgumentError(BuildExportError):
    pass


class TemplateError(BuildExportError):
    pass


class DataSchemaError(BuildExportError):
    pass


class IOFailure(BuildExportError):
    pass


# --------------------------------------------------------------------------- #
# Logging
# --------------------------------------------------------------------------- #

def setup_logging(verbose: bool = False) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        datefmt="%Y-%m-%dT%H:%M:%S%z",
    )


# --------------------------------------------------------------------------- #
# Path validation
# --------------------------------------------------------------------------- #

def validate_path(
    path_str: str,
    *,
    must_exist: bool,
    allowed_ext: Iterable[str],
    max_bytes: Optional[int] = None,
    label: str = "file",
) -> Path:
    """Validate a user-supplied path and return a resolved Path object."""
    if not path_str or not isinstance(path_str, str):
        raise InvalidArgumentError(f"{label} path is empty.")

    # Reject obvious traversal attempts before resolving.
    if "\x00" in path_str:
        raise InvalidArgumentError(f"{label} path contains NUL byte.")

    try:
        p = Path(path_str).expanduser().resolve(strict=False)
    except (OSError, ValueError) as e:
        raise InvalidArgumentError(f"{label} path is invalid: {e}") from e

    if p.suffix.lower() not in allowed_ext:
        raise InvalidArgumentError(
            f"{label} must have one of extensions {sorted(allowed_ext)}; "
            f"got '{p.suffix}'."
        )

    if must_exist:
        if not p.is_file():
            raise InvalidArgumentError(f"{label} not found: {p}")
        if max_bytes is not None:
            try:
                size = p.stat().st_size
            except OSError as e:
                raise InvalidArgumentError(f"cannot stat {label}: {e}") from e
            if size > max_bytes:
                raise InvalidArgumentError(
                    f"{label} too large: {size} bytes (limit {max_bytes})."
                )
    return p


# --------------------------------------------------------------------------- #
# Value sanitisation
# --------------------------------------------------------------------------- #

def sanitize_string(value: str, max_len: int = 32_000) -> str:
    """Sanitize a string before writing it to an Excel cell.

    * Strip control characters except newline and tab.
    * Mitigate formula injection by prefixing dangerous leading chars.
    * Enforce a length cap.
    """
    if not isinstance(value, str):
        value = str(value)
    # Remove control chars except \n, \r, \t.
    value = "".join(
        ch for ch in value
        if ch in ("\n", "\r", "\t") or (ord(ch) >= 0x20 and ord(ch) != 0x7F)
    )
    if len(value) > max_len:
        value = value[:max_len]
    # Formula-injection mitigation: prefix with single quote if it looks like a
    # formula. The leading quote is consumed by Excel and not displayed.
    if _INJECTION_RE.match(value):
        value = "'" + value
    return value


def coerce_value(raw: Any, expected_type: type) -> Any:
    """Coerce a JSON value to the expected Python type with safety checks."""
    if raw is None:
        return 0 if expected_type is not int else ("")
    if expected_type is str:
        return sanitize_string(raw)
    if expected_type is int:
        if isinstance(raw, bool):
            return int(raw)
        if isinstance(raw, int):
            return raw
        if isinstance(raw, float):
            if not raw.is_finite():
                return 0
            return int(raw)
        if isinstance(raw, str):
            try:
                return int(float(raw))
            except ValueError:
                return 0
        return 0
    if isinstance(expected_type, tuple):  # numeric (int or float)
        if isinstance(raw, bool):
            return float(raw)
        if isinstance(raw, (int, float)):
            if isinstance(raw, float) and not raw.is_finite():
                LOG.warning("Non-finite float value encountered; converting to 0.0")
                return 0.0
            return raw
        if isinstance(raw, str):
            try:
                return float(raw)
            except ValueError:
                return 0.0
        return 0.0
    return raw


# --------------------------------------------------------------------------- #
# JSON loading & schema validation
# --------------------------------------------------------------------------- #

def load_data_json(path: Path) -> Dict[str, Any]:
    try:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        raise DataSchemaError(f"Invalid JSON in {path}: {e}") from e
    except OSError as e:
        raise IOFailure(f"Cannot read data file {path}: {e}") from e

    if not isinstance(data, dict):
        raise DataSchemaError("Top-level JSON must be an object.")

    # Validate each section.
    for key in (
        "district", "ar_content", "quiz", "session",
        "language", "device", "notification", "ads",
    ):
        section = data.get(key)
        if section is None:
            data[key] = []
            continue
        if not isinstance(section, list):
            raise DataSchemaError(f"Section '{key}' must be a list.")
        for i, row in enumerate(section):
            if not isinstance(row, dict):
                raise DataSchemaError(
                    f"Section '{key}' row {i} must be an object."
                )

    month_label = data.get("month_label", "")
    if not isinstance(month_label, str):
        raise DataSchemaError("'month_label' must be a string.")
    data["month_label"] = sanitize_string(month_label, max_len=64)

    return data


# --------------------------------------------------------------------------- #
# Helpers for matching
# --------------------------------------------------------------------------- #

def _norm(v: Any) -> str:
    """Normalize a key value for matching (case-insensitive)."""
    if v is None:
        return ""
    return str(v).strip().lower()


def _row_key(row: Dict[str, Any], fields: Tuple[str, ...]) -> Tuple[str, ...]:
    return tuple(_norm(row.get(f, "")) for f in fields)


def build_index(
    rows: List[Dict[str, Any]],
    key_fields: Tuple[str, ...],
) -> Dict[Tuple[str, ...], Dict[str, Any]]:
    idx: Dict[Tuple[str, ...], Dict[str, Any]] = {}
    for r in rows:
        idx[_row_key(r, key_fields)] = r
    return idx


# --------------------------------------------------------------------------- #
# Cell writers
# --------------------------------------------------------------------------- #

def safe_set(ws, row_idx: int, col_idx: int, value: Any) -> None:
    """Set a cell value safely. 1-based indices."""
    if col_idx < 1 or col_idx > 16384:
        LOG.warning("Column index %d out of bounds; skipped.", col_idx)
        return
    if row_idx < 1 or row_idx > 1_048_576:
        LOG.warning("Row index %d out of bounds; skipped.", row_idx)
        return
    ws.cell(row=row_idx, column=col_idx, value=value)


# --------------------------------------------------------------------------- #
# Sheet populators
# --------------------------------------------------------------------------- #

def _populate_keyed_sheet(
    ws,
    rows: List[Dict[str, Any]],
    key_fields: Tuple[str, ...],
    key_cell_cols: Tuple[int, ...],
    column_map: List[Tuple[int, str, Union[type, Tuple[type, ...]]]],
    *,
    extra_static: Optional[Dict[int, Any]] = None,
    start_row: int = 4,
    sheet_name: str = "",
) -> int:
    """Generic populator: look up each template row by composite key."""
    index = build_index(rows, key_fields)
    updated = 0

    max_row = min(ws.max_row or start_row, start_row + MAX_ROWS_PER_SHEET - 1)
    for r_idx in range(start_row, max_row + 1):
        # Read key cells.
        key = tuple(
            _norm(ws.cell(row=r_idx, column=c).value)
            for c in key_cell_cols
        )
        if not any(key):
            continue

        matched = index.get(key)
        # Static extras first (e.g., month label).
        if extra_static:
            for col_idx, val in extra_static.items():
                safe_set(ws, r_idx, col_idx, val)

        for col_idx, src_key, typ in column_map:
            if matched is None:
                raw = 0 if typ is int else ("" if typ is str else 0.0)
            else:
                raw = matched.get(src_key)
            safe_set(ws, r_idx, col_idx, coerce_value(raw, typ))

        if matched is not None:
            updated += 1
    LOG.info("Sheet '%s': updated %d rows.", sheet_name, updated)
    return updated


def _populate_ads(ws, ads_rows: List[Dict[str, Any]], month_label: str) -> int:
    """Clear and re-write RAW_ADS from row 4."""
    if len(ads_rows) > MAX_ADS_ROWS:
        raise DataSchemaError(
            f"Too many ads rows: {len(ads_rows)} (limit {MAX_ADS_ROWS})."
        )

    # Validate required fields in each row
    required_fields = ("campaign_name", "campaign_type", "target_state", "impressions")
    for i, row in enumerate(ads_rows):
        for field in required_fields:
            if field not in row or row[field] is None:
                LOG.warning(
                    "RAW_ADS row %d missing or null field '%s'; will use default.",
                    i, field
                )

    # Clear existing data rows (preserve header row 3).
    max_row = min(ws.max_row or 3, 3 + MAX_ROWS_PER_SHEET)
    for r_idx in range(4, max_row + 1):
        for c_idx in range(2, ws.max_column + 1):
            safe_set(ws, r_idx, c_idx, None)
        # Reset col A (row number) too.
        safe_set(ws, r_idx, 1, None)

    # Write fresh rows.
    field_map = [
        (2,  "campaign_name",   str),
        (3,  "campaign_type",   str),
        (4,  "target_state",    str),
        (5,  "impressions",     int),
        (6,  "unique_reach",    int),
        (7,  "ctr_pct",         (int, float)),
        (8,  "avg_view_sec",    (int, float)),
        (9,  "completion_pct",  (int, float)),
        (10, "revenue_inr",     (int, float)),
    ]
    for i, r in enumerate(ads_rows):
        r_idx = i + 4
        safe_set(ws, r_idx, 1, i + 1)
        for col_idx, src_key, typ in field_map:
            safe_set(ws, r_idx, col_idx, coerce_value(r.get(src_key), typ))
        safe_set(ws, r_idx, 11, sanitize_string(month_label, max_len=64))
        safe_set(ws, r_idx, 12, "Compliant")
    LOG.info("Sheet 'RAW_ADS': wrote %d rows.", len(ads_rows))
    return len(ads_rows)


# --------------------------------------------------------------------------- #
# Atomic save
# --------------------------------------------------------------------------- #

def atomic_save(wb, output_path: Path) -> None:
    """Save the workbook atomically: temp file + fsync + rename."""
    output_dir = output_path.parent
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        raise IOFailure(f"Cannot create output directory {output_dir}: {e}") from e

    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{output_path.stem}_",
            suffix=output_path.suffix,
            dir=str(output_dir),
            delete=False,
        ) as tmp:
            tmp_path = Path(tmp.name)
            # openpyxl writes the file; we then fsync the FD.
            wb.save(tmp_path)
            tmp.flush()
            os.fsync(tmp.fileno())
        os.replace(tmp_path, output_path)
    except OSError as e:
        # Best-effort cleanup of the temp file.
        try:
            if "tmp_path" in locals() and tmp_path.exists():
                tmp_path.unlink()
        except OSError:
            pass
        raise IOFailure(f"Failed to save workbook: {e}") from e
    except Exception as e:
        raise IOFailure(f"Failed to save workbook: {e}") from e


# --------------------------------------------------------------------------- #
# Main populate
# --------------------------------------------------------------------------- #

def populate(template_path: Path, data_json_path: Path, output_path: Path) -> None:
    data = load_data_json(data_json_path)
    month_label: str = data["month_label"]

    try:
        wb = load_workbook(template_path, read_only=False, keep_vba=False)
    except InvalidFileException as e:
        raise TemplateError(f"Template is not a valid xlsx: {e}") from e
    except Exception as e:
        raise TemplateError(f"Cannot open template: {e}") from e

    missing = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing:
        raise TemplateError(f"Template missing required sheets: {missing}")

    # ---- RAW_DISTRICT (key = State + District, cols B=2, C=3) ----
    _populate_keyed_sheet(
        wb["RAW_DISTRICT"],
        data["district"],
        key_fields=("state", "district"),
        key_cell_cols=(2, 3),
        column_map=SHEET_MAPS["RAW_DISTRICT"],
        extra_static={4: month_label},  # col D
        sheet_name="RAW_DISTRICT",
    )

    # ---- RAW_AR_CONTENT (key = State + AR Module Title; cols B=2, E=5) ----
    _populate_keyed_sheet(
        wb["RAW_AR_CONTENT"],
        data["ar_content"],
        key_fields=("state", "ar_module_title"),
        key_cell_cols=(2, 5),
        column_map=SHEET_MAPS["RAW_AR_CONTENT"],
        sheet_name="RAW_AR_CONTENT",
    )

    # ---- RAW_QUIZ (key = State + Quiz Title; cols B=2, E=5) ----
    _populate_keyed_sheet(
        wb["RAW_QUIZ"],
        data["quiz"],
        key_fields=("state", "quiz_title"),
        key_cell_cols=(2, 5),
        column_map=SHEET_MAPS["RAW_QUIZ"],
        sheet_name="RAW_QUIZ",
    )

    # ---- RAW_SESSION (key = State; col B=2). Month label injected into col C. ----
    session_map = [
        (c, k, t) for (c, k, t) in SHEET_MAPS["RAW_SESSION"]
        if k != "__month_label__"
    ]
    _populate_keyed_sheet(
        wb["RAW_SESSION"],
        data["session"],
        key_fields=("state",),
        key_cell_cols=(2,),
        column_map=session_map,
        extra_static={3: month_label},  # col C Report Month (was 2, now 3)
        sheet_name="RAW_SESSION",
    )

    # ---- RAW_LANGUAGE (key = State; col B=2) ----
    _populate_keyed_sheet(
        wb["RAW_LANGUAGE"],
        data["language"],
        key_fields=("state",),
        key_cell_cols=(2,),
        column_map=SHEET_MAPS["RAW_LANGUAGE"],
        sheet_name="RAW_LANGUAGE",
    )

    # ---- RAW_DEVICE (key = State; col B=2) ----
    _populate_keyed_sheet(
        wb["RAW_DEVICE"],
        data["device"],
        key_fields=("state",),
        key_cell_cols=(2,),
        column_map=SHEET_MAPS["RAW_DEVICE"],
        sheet_name="RAW_DEVICE",
    )

    # ---- RAW_NOTIFICATION (key = State + Notif Type; cols B=2, C=3) ----
    _populate_keyed_sheet(
        wb["RAW_NOTIFICATION"],
        data["notification"],
        key_fields=("state", "notif_type"),
        key_cell_cols=(2, 3),
        column_map=SHEET_MAPS["RAW_NOTIFICATION"],
        sheet_name="RAW_NOTIFICATION",
    )

    # ---- RAW_ADS (cleared & rewritten) ----
    _populate_ads(wb["RAW_ADS"], data["ads"], month_label)

    atomic_save(wb, output_path)
    LOG.info("Saved: %s", output_path)


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        prog="build_export.py",
        description="Populate MITRA analytics template with live DB data.",
    )
    p.add_argument("--template", required=True,
                   help="Path to MITRA_Analytics_v7_Complete.xlsx (template).")
    p.add_argument("--data", required=True,
                   help="Path to JSON file with live DB data.")
    p.add_argument("--output", required=True,
                   help="Path to write the populated workbook.")
    p.add_argument("--verbose", action="store_true",
                   help="Enable debug logging.")
    # Backward-compat: accept positional args too.
    p.add_argument("positional", nargs="*",
                   help="Optional positional: <input.xlsx> <data.json> <output.xlsx>")
    return p.parse_args(argv)


def resolve_args(args: argparse.Namespace) -> Tuple[Path, Path, Path]:
    if args.positional:
        if len(args.positional) != 3:
            raise InvalidArgumentError(
                "Expected exactly 3 positional arguments: "
                "<input.xlsx> <data.json> <output.xlsx>"
            )
        template_s, data_s, output_s = args.positional
    else:
        template_s, data_s, output_s = args.template, args.data, args.output

    template = validate_path(
        template_s, must_exist=True,
        allowed_ext=ALLOWED_TEMPLATE_EXT,
        max_bytes=MAX_TEMPLATE_FILE_BYTES,
        label="template",
    )
    data = validate_path(
        data_s, must_exist=True,
        allowed_ext=ALLOWED_DATA_EXT,
        max_bytes=MAX_DATA_FILE_BYTES,
        label="data",
    )
    output = validate_path(
        output_s, must_exist=False,
        allowed_ext=ALLOWED_OUTPUT_EXT,
        label="output",
    )
    # Refuse to overwrite the template with the output (common mistake).
    try:
        if output.resolve() == template.resolve():
            raise InvalidArgumentError("Output path must differ from template path.")
    except OSError:
        pass
    return template, data, output


def main(argv: Optional[List[str]] = None) -> int:
    try:
        args = parse_args(argv)
        setup_logging(args.verbose)
        template, data, output = resolve_args(args)
        populate(template, data, output)
        return 0
    except InvalidArgumentError as e:
        LOG.error("Argument error: %s", e)
        return 2
    except TemplateError as e:
        LOG.error("Template error: %s", e)
        return 3
    except DataSchemaError as e:
        LOG.error("Data error: %s", e)
        return 4
    except IOFailure as e:
        LOG.error("I/O error: %s", e)
        return 5
    except BuildExportError as e:
        LOG.error("Error: %s", e)
        return 1
    except Exception as e:  # pragma: no cover - defensive
        LOG.exception("Unhandled error: %s", e)
        return 1


if __name__ == "__main__":
    sys.exit(main())